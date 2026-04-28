#!/usr/bin/env python3
"""
Portfolio Data Manager
======================
- json_to_excel : Reads portfolio.json → writes portfolio_editor.xlsx (human-friendly editor)
- excel_to_json : Reads portfolio_editor.xlsx → writes portfolio.json (for the website)

Usage:
    python data_manager.py json_to_excel
    python data_manager.py excel_to_json
"""

import json, sys, os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(BASE_DIR, "data", "portfolio.json")
XLSX_PATH = os.path.join(BASE_DIR, "data", "portfolio_editor.xlsx")

# ── Color palette ─────────────────────────────────────────────────────────────
NAVY   = "1A2E4A"
TEAL   = "0E7C7B"
SILVER = "E8EFF5"
WHITE  = "FFFFFF"
ACCENT = "F0A500"

def hdr(ws, row, col, text, bg=NAVY, fg=WHITE, size=11, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    return c

def cell(ws, row, col, value="", wrap=True, bold=False, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=10, bold=bold)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    return c

# ══════════════════════════════════════════════════════════════════════════════
#  JSON → EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def json_to_excel():
    with open(JSON_PATH, encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet: PROFILE ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Profile")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55

    hdr(ws, 1, 1, "PROFILE", bg=NAVY, size=13)
    hdr(ws, 1, 2, "Edit values in column B", bg=TEAL, size=11)

    fields = ["name","title","tagline","location","phone","email",
              "photo","about","linkedin","github","website"]
    p = data["profile"]
    for i, f in enumerate(fields, start=2):
        c = ws.cell(row=i, column=1, value=f)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=SILVER)
        cell(ws, i, 2, p.get(f,""))
        ws.row_dimensions[i].height = 20 if f != "about" else 50

    # ── Sheet: SKILLS ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Skills")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 60
    hdr(ws, 1, 1, "Category"); hdr(ws, 1, 2, "Skills (comma separated)")
    for i, sk in enumerate(data["skills"], start=2):
        cell(ws, i, 1, sk["category"], bold=True, bg=SILVER)
        cell(ws, i, 2, ", ".join(sk["items"]))

    # ── Sheet: EXPERIENCE ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Experience")
    cols = ["id","role","company","location","start","end","bullets (one per line)"]
    widths = [10, 30, 20, 20, 12, 12, 70]
    for ci, (col, w) in enumerate(zip(cols, widths), start=1):
        hdr(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for i, exp in enumerate(data["experience"], start=2):
        bg = SILVER if i % 2 == 0 else WHITE
        cell(ws, i, 1, exp["id"], bg=bg)
        cell(ws, i, 2, exp["role"], bold=True, bg=bg)
        cell(ws, i, 3, exp["company"], bg=bg)
        cell(ws, i, 4, exp["location"], bg=bg)
        cell(ws, i, 5, exp["start"], bg=bg)
        cell(ws, i, 6, exp["end"], bg=bg)
        cell(ws, i, 7, "\n".join(exp["bullets"]), bg=bg)
        ws.row_dimensions[i].height = max(40, 15 * len(exp["bullets"]))

    # ── Sheet: EDUCATION ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Education")
    for ci, col in enumerate(["degree","institution","start","end"], start=1):
        hdr(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = [40,30,10,10][ci-1]
    for i, ed in enumerate(data["education"], start=2):
        bg = SILVER if i % 2 == 0 else WHITE
        cell(ws, i, 1, ed["degree"], bold=True, bg=bg)
        cell(ws, i, 2, ed.get("institution",""), bg=bg)
        cell(ws, i, 3, ed["start"], bg=bg)
        cell(ws, i, 4, ed["end"], bg=bg)

    # ── Sheet: CERTIFICATIONS ─────────────────────────────────────────────────
    ws = wb.create_sheet("Certifications")
    for ci, col in enumerate(["name","issuer","year"], start=1):
        hdr(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = [45,30,12][ci-1]
    for i, cert in enumerate(data["certifications"], start=2):
        bg = SILVER if i % 2 == 0 else WHITE
        cell(ws, i, 1, cert["name"], bold=True, bg=bg)
        cell(ws, i, 2, cert.get("issuer",""), bg=bg)
        cell(ws, i, 3, cert.get("year",""), bg=bg)

    # ── Sheet: LANGUAGES ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Languages")
    for ci, col in enumerate(["language","level"], start=1):
        hdr(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = [25,25][ci-1]
    for i, lang in enumerate(data["languages"], start=2):
        bg = SILVER if i % 2 == 0 else WHITE
        cell(ws, i, 1, lang["language"], bg=bg)
        cell(ws, i, 2, lang["level"], bg=bg)

    # ── Sheet: PROJECTS ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Projects")
    proj_cols = ["id","title","category","tags (comma sep)","description",
                 "images (filenames, comma sep)","embed (URL/iframe src)",
                 "link","year","featured (TRUE/FALSE)"]
    proj_widths = [10, 35, 18, 35, 55, 35, 35, 25, 8, 16]
    for ci, (col, w) in enumerate(zip(proj_cols, proj_widths), start=1):
        hdr(ws, 1, ci, col)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 30
    note = ws.cell(row=2, column=1,
                   value="⬇ Add new rows below. Do NOT change IDs of existing projects (breaks image links). Leave ID blank for new ones — it will be auto-generated.")
    note.font = Font(italic=True, color="888888", size=9)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    for i, proj in enumerate(data["projects"], start=3):
        bg = SILVER if i % 2 == 1 else WHITE
        vals = [
            proj["id"],
            proj["title"],
            proj["category"],
            ", ".join(proj.get("tags",[])),
            proj["description"],
            ", ".join(proj.get("images",[])),
            proj.get("embed",""),
            proj.get("link",""),
            str(proj.get("year","")),
            str(proj.get("featured", False)).upper()
        ]
        for ci, v in enumerate(vals, start=1):
            cell(ws, i, ci, v, bg=bg)
        ws.row_dimensions[i].height = 50

    # ── Sheet: HOW TO USE ─────────────────────────────────────────────────────
    ws = wb.create_sheet("HOW TO USE", 0)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70
    hdr(ws, 1, 1, "PORTFOLIO EDITOR — HOW TO USE", bg=TEAL, size=14)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 35

    instructions = [
        ("Step 1", "Edit any sheet (Profile, Skills, Experience, Education, Certifications, Languages, Projects)."),
        ("Step 2", "Save this file (Ctrl+S / Cmd+S)."),
        ("Step 3", "Run:  python data_manager.py excel_to_json  — this regenerates portfolio.json."),
        ("Step 4", "Open index.html in your browser. Changes appear instantly."),
        ("Adding Projects", "Go to Projects sheet. Add a new row at the bottom. Leave ID blank — it will be auto-generated."),
        ("Project Images", "Put image files in the  assets/  folder. In the 'images' column write filenames e.g.: dashboard.png, chart.jpg"),
        ("Embed Content", "Paste a Power BI embed URL, YouTube embed URL, or any iframe src into the 'embed' column."),
        ("Featured Projects", "Set 'featured' to TRUE to show the project prominently on the homepage."),
        ("Bullets", "In the Experience sheet, put each bullet on its own line inside the cell (Alt+Enter for new line in Excel)."),
        ("Tags", "Separate multiple tags with commas: Power BI, DAX, Python"),
    ]
    for i, (step, desc) in enumerate(instructions, start=2):
        c1 = ws.cell(row=i, column=1, value=step)
        c1.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c1.fill = PatternFill("solid", fgColor=NAVY)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2 = ws.cell(row=i, column=2, value=desc)
        c2.font = Font(name="Calibri", size=10)
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        bg = SILVER if i % 2 == 0 else WHITE
        c2.fill = PatternFill("solid", fgColor=bg)
        ws.row_dimensions[i].height = 22

    wb.save(XLSX_PATH)
    print(f"✅ Excel created: {XLSX_PATH}")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL → JSON
# ══════════════════════════════════════════════════════════════════════════════
def excel_to_json():
    wb = load_workbook(XLSX_PATH, data_only=True)
    data = {}

    # Profile
    ws = wb["Profile"]
    profile = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            profile[row[0]] = row[1] or ""
    data["profile"] = profile

    # Skills
    ws = wb["Skills"]
    skills = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            items = [x.strip() for x in str(row[1] or "").split(",") if x.strip()]
            skills.append({"category": row[0], "items": items})
    data["skills"] = skills

    # Experience
    ws = wb["Experience"]
    experience = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            bullets = [b.strip() for b in str(row[6] or "").split("\n") if b.strip()]
            experience.append({
                "id": row[0] or f"exp{len(experience)+1}",
                "role": row[1] or "",
                "company": row[2] or "",
                "location": row[3] or "",
                "start": str(row[4] or ""),
                "end": str(row[5] or ""),
                "bullets": bullets
            })
    data["experience"] = experience

    # Education
    ws = wb["Education"]
    education = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            education.append({
                "degree": row[0] or "",
                "institution": row[1] or "",
                "start": str(row[2] or ""),
                "end": str(row[3] or "")
            })
    data["education"] = education

    # Certifications
    ws = wb["Certifications"]
    certifications = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            certifications.append({
                "name": row[0] or "",
                "issuer": row[1] or "",
                "year": str(row[2] or "")
            })
    data["certifications"] = certifications

    # Languages
    ws = wb["Languages"]
    languages = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            languages.append({"language": row[0], "level": row[1] or ""})
    data["languages"] = languages

    # Projects
    ws = wb["Projects"]
    projects = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        proj_id = row[0] or f"proj{len(projects)+1}"
        tags = [t.strip() for t in str(row[3] or "").split(",") if t.strip()]
        images = [x.strip() for x in str(row[5] or "").split(",") if x.strip()]
        featured = str(row[9] or "").strip().upper() == "TRUE"
        projects.append({
            "id": proj_id,
            "title": row[1] or "",
            "category": row[2] or "",
            "tags": tags,
            "description": row[4] or "",
            "images": images,
            "embed": row[6] or "",
            "link": row[7] or "",
            "year": str(row[8] or ""),
            "featured": featured
        })
    data["projects"] = projects

    from datetime import date
    data["meta"] = {"last_updated": str(date.today()), "version": "1.0"}

    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON updated: {JSON_PATH}")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""
    if cmd == "json_to_excel":
        json_to_excel()
    elif cmd == "excel_to_json":
        excel_to_json()
    else:
        print("Usage: python data_manager.py [json_to_excel | excel_to_json]")
